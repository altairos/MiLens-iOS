#!/usr/bin/env python3
"""MiLens 本地化（String Catalog）导出 / 导入 / 校验工具。

支持任意语言的 String Catalog（.xcstrings）与 Excel（.xlsx）互转，
便于外部翻译人员离线翻译后再回写。结构按 Apple String Catalog 规范组织（含 variations/plural 复数条目：
导出时每个复数变体占一行、variation 列标注变体名，导入时按 (key, variation)
合并回 variations.plural），新增语言时无需改动本工具——只需在 project.yml
的 knownRegions 追加语言代码，再用本工具导出该语言的空列即可。

依赖：openpyxl（pip install -r tools/requirements.txt）。

典型工作流：
  # 1. 导出（生成待翻译 Excel，en 列留空待填）
  python tools/localization.py export \\
      MiLens/Resources/Localizable.xcstrings \\
      MiLens/Resources/InfoPlist.xcstrings \\
      build/localization.xlsx --lang en

  # 2. 翻译人员填写 Excel 中的 en 列

  # 3. 导入回写（自动在 .xcstrings 创建 en locale）
  python tools/localization.py import build/localization.xlsx \\
      MiLens/Resources/Localizable.xcstrings --lang en

  # 4. 校验（JSON 合法性 + 代码引用一致性 + 缺译检测）
  python tools/localization.py check \\
      MiLens/Resources/Localizable.xcstrings \\
      MiLens/Resources/InfoPlist.xcstrings \\
      --project-yml project.yml --source-root MiLens

详见 DEVELOPMENT.md §4.4。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # 延迟到子命令真正需要 openpyxl 时才报错
    Workbook = None  # type: ignore[assignment,misc]
    load_workbook = None  # type: ignore[assignment]


# --------------------------------------------------------------------------- #
# xcstrings 读写
# --------------------------------------------------------------------------- #

def load_xcstrings(path: Path) -> dict:
    """读取并解析 .xcstrings（JSON）。"""
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def dumps_xcstrings(obj: dict) -> str:
    """Xcode 风格序列化：2 空格缩进、键字母序、冒号后带空格、不转义非 ASCII。

    Apple String Catalog 在键与冒号间留一个空格（``"key" : value``），
    ``json.dumps`` 默认无此空格。下方正则仅匹配「字符串字面量 + 冒号」——
    JSON 中 value 字符串后不会紧跟冒号，故替换只作用于 key。
    """
    text = json.dumps(obj, indent=2, ensure_ascii=False, sort_keys=True)
    text = re.sub(r'("(?:\\.|[^"\\])*")(\s*):', r"\1 :", text)
    return text + "\n"


def save_xcstrings(path: Path, obj: dict) -> None:
    """以 Xcode 风格写回 .xcstrings。"""
    path.write_text(dumps_xcstrings(obj), encoding="utf-8")


# --------------------------------------------------------------------------- #
# 语言列管理
# --------------------------------------------------------------------------- #

def collect_languages(obj: dict) -> list[str]:
    """收集 xcstrings 中出现过的所有语言代码（含 sourceLanguage，置于首位）。"""
    langs: list[str] = []
    seen: set[str] = set()
    source = obj.get("sourceLanguage", "")
    if source:
        langs.append(source)
        seen.add(source)
    for entry in obj.get("strings", {}).values():
        for lang in entry.get("localizations", {}).keys():
            if lang not in seen:
                seen.add(lang)
                langs.append(lang)
    return langs


# 语言代码 -> 展示名（CLI 统计表与 GUI 进度面板共用）
LANG_NAMES: dict[str, str] = {
    "zh-Hans": "简体中文",
    "zh-Hant": "繁體中文",
    "ja": "日本語",
    "ko": "한국어",
    "en": "English",
    "fr": "Français",
    "de": "Deutsch",
}


@dataclass
class LangStatus:
    """某语言的翻译进度统计（CLI check 与 GUI 共用）。"""

    lang: str
    name: str
    total: int = 0
    translated: int = 0
    needs_review: int = 0
    missing: int = 0
    is_source: bool = False

    @property
    def pct(self) -> int:
        if self.total == 0:
            return 0
        # 已译 + 初译待审 都算"有译文"，计入完成度
        return round((self.translated + self.needs_review) * 100 / self.total)


def scan_statuses(objs: list[tuple[Path, dict]], known: list[str]) -> list[LangStatus]:
    """统计各语言翻译进度（合并多个已加载的 .xcstrings）。

    缺译 = 无译文或 state=="new"；初译待审 = state=="needs_review"；
    已译 = state=="translated"。源语言行仅作参照（translated=total）。
    """
    source = ""
    stats: dict[str, LangStatus] = {}
    for _path, obj in objs:
        if not source:
            source = obj.get("sourceLanguage", "")
        strings = obj.get("strings", {})
        for lang in known:
            st = stats.setdefault(lang, LangStatus(lang, LANG_NAMES.get(lang, lang)))
            if lang == source:
                st.is_source = True
                st.total += len(strings)
                st.translated += len(strings)
                continue
            for key in strings:
                value, state = entry_lang_status(strings[key], lang)
                st.total += 1
                if not value or state == "new":
                    st.missing += 1
                elif state == "needs_review":
                    st.needs_review += 1
                else:
                    st.translated += 1
    return [stats[l] for l in known if l in stats]


def entry_value_state(entry: dict, lang: str) -> tuple[str, str]:
    """返回 (value, state)；无 stringUnit 或无该语言时返回 ("", "")。

    复数条目没有 stringUnit（值在 variations.plural），需用
    entry_lang_status / entry_lang_rows 处理。
    """
    loc = entry.get("localizations", {}).get(lang, {})
    unit = loc.get("stringUnit")
    if not unit:
        return "", ""
    return unit.get("value") or "", unit.get("state") or ""


# 复数变体固定顺序（Apple CLDR 顺序）
PLURAL_VARIATION_ORDER = ["zero", "one", "two", "few", "many", "other"]

# 需要 one/other 双变体的语言（其余语言的复数形态翻译时合并为单条变体即可）
PLURAL_DUAL_LANGS = {"en", "de", "fr"}

# 格式占位符：%d、%@、%lld、%1$@、%.2f 等（%% 为转义字面百分号，不匹配；
# 用非捕获组避免 findall 返回捕获组而非完整匹配）
_PLACEHOLDER_RE = re.compile(r"%(?:\d+\$)?[-+0-9.]*[a-zA-Z@]+")


def entry_plural(entry: dict, lang: str) -> dict | None:
    """返回该语言的复数变体 {variation: (value, state)}；非复数条目返回 None。"""
    plural = entry.get("localizations", {}).get(lang, {}).get("variations", {}).get("plural")
    if not plural:
        return None
    out: dict[str, tuple[str, str]] = {}
    for var, node in plural.items():
        if var == "pluralRuleType":
            continue
        unit = node.get("stringUnit") if isinstance(node, dict) else None
        out[var] = (unit.get("value") or "", unit.get("state") or "") if unit else ("", "")
    return out or None


def entry_lang_rows(entry: dict, lang: str) -> list[tuple[str, str, str]]:
    """展开条目在该语言的取值行 [(variation, value, state)]。

    普通字符串条目返回单行（variation 为空串）；复数条目每个变体一行。
    """
    plural = entry_plural(entry, lang)
    if plural:
        return [(var, val, st) for var, (val, st) in plural.items()]
    value, state = entry_value_state(entry, lang)
    return [("", value, state)]


def entry_lang_status(entry: dict, lang: str) -> tuple[str, str]:
    """聚合视图 (value, state)，供缺译检测与进度统计使用。

    复数条目：value 为各变体值以换行连接；任一变体缺失/空值/state=="new"
    时整体记为 "new"（视为缺译），否则取 needs_review/translated 中更保守者。
    """
    plural = entry_plural(entry, lang)
    if not plural:
        return entry_value_state(entry, lang)
    values: list[str] = []
    worst = "translated"
    # 变体并集 = 语言要求变体 + 实际存在变体（缺 required 变体整体视为缺译）
    seen: set[str] = set()
    all_vars: list[str] = []
    for var in list(plural_required_variations(lang)) + list(plural.keys()):
        if var not in seen:
            seen.add(var)
            all_vars.append(var)
    for var in all_vars:
        value, state = plural.get(var, ("", ""))
        if not value or state == "new":
            worst = "new"
        elif state == "needs_review" and worst != "new":
            worst = "needs_review"
        if value:
            values.append(value)
    return "\n".join(values), worst


def plural_required_variations(lang: str) -> tuple[str, ...]:
    """该语言要求的复数变体集合：en/de/fr 需 one+other，其余语言单条 other。"""
    if lang in PLURAL_DUAL_LANGS:
        return ("one", "other")
    return ("other",)


def plural_problems(path: Path, key: str, entry: dict, lang: str, source: str
                    ) -> list[tuple[str, str, str]]:
    """复数条目完整性问题 [(lang, key, note)]：缺变体 + 占位符漂移。"""
    plural = entry_plural(entry, lang)
    if plural is None:
        return []
    problems: list[tuple[str, str, str]] = []
    for var in plural_required_variations(lang):
        if var not in plural:
            problems.append((lang, key, f"{path.stem}：复数缺变体「{var}」"))
        else:
            value, state = plural[var]
            if not value or state == "new":
                problems.append((lang, key, f"{path.stem}：复数变体「{var}」未完成（state={state!r}）"))
    # 占位符漂移：主变体（other）必须完整保留源语言占位符（防漏 %d/%@）；
    # 其余变体（如 one）允许按语言习惯写死单数形式（Apple 推荐“1 day”写法）
    src_phs: set[str] = set()
    src_plural = entry_plural(entry, source)
    if src_plural is None:
        src_value, _ = entry_value_state(entry, source)
        src_phs = set(_PLACEHOLDER_RE.findall(src_value))
    else:
        for src_value, _ in src_plural.values():
            src_phs |= set(_PLACEHOLDER_RE.findall(src_value))
    if src_phs:
        main_var = "other" if "other" in plural else next(iter(plural))
        main_phs = set(_PLACEHOLDER_RE.findall(plural[main_var][0]))
        if main_phs != src_phs:
            problems.append((lang, key, f"{path.stem}：复数占位符与源语言不一致"
                             f"（源 {sorted(src_phs)} vs 主变体「{main_var}」{sorted(main_phs)}）"))
    return problems


def missing_problems(obj: dict, path: Path, known: list[str]
                     ) -> list[tuple[str, str, str]]:
    """缺译 + 复数完整性问题清单 [(lang, key, note)]；check 与 GUI 共用同语义。"""
    source = obj.get("sourceLanguage", "")
    problems: list[tuple[str, str, str]] = []
    for key in sorted(obj.get("strings", {}).keys()):
        entry = obj["strings"][key]
        for lang in known:
            if lang == source:
                continue
            if entry_plural(entry, lang) is not None:
                # 复数条目：变体级检查（缺变体 > 变体未完成 > 占位符漂移），取首条保持 1 对 1
                detail = plural_problems(path, key, entry, lang, source)
                if detail:
                    problems.append(detail[0])
                continue
            value, state = entry_lang_status(entry, lang)
            if not value or state == "new":
                note = f"{path.stem}：未完成（state={state!r}）" if state else f"{path.stem}：无译文"
                problems.append((lang, key, note))
    return problems


def review_problems(obj: dict, path: Path, known: list[str]
                    ) -> list[tuple[str, str, str]]:
    """初译待审（needs_review）问题清单 [(lang, key, note)]。

    与 missing_problems 互补：缺译（new/空）由 missing_problems 报；
    本函数只报 value 非空但 state=="needs_review" 的条目，供发布门禁
    （check --strict）阻断"未审校译文"上线。
    """
    source = obj.get("sourceLanguage", "")
    problems: list[tuple[str, str, str]] = []
    for key in sorted(obj.get("strings", {}).keys()):
        entry = obj["strings"][key]
        for lang in known:
            if lang == source:
                continue
            value, worst = entry_lang_status(entry, lang)
            # value 非空 + worst==needs_review（复数取聚合 worst）
            if value and worst == "needs_review":
                problems.append((lang, key, f"{path.stem}：初译待审（needs_review）"))
    return problems


def placeholder_problems(obj: dict, path: Path, known: list[str]
                         ) -> list[tuple[str, str, str]]:
    """普通条目占位符漂移清单 [(lang, key, note)]。

    复数条目的占位符漂移由 plural_problems 检查；本函数覆盖普通 stringUnit
    条目：译文占位符集合必须与源语言一致（防漏 %d/%@ 等运行时插值）。
    缺译（无 value）由 missing_problems 负责，此处跳过。
    """
    source = obj.get("sourceLanguage", "")
    problems: list[tuple[str, str, str]] = []
    for key in sorted(obj.get("strings", {}).keys()):
        entry = obj["strings"][key]
        src_value, _ = entry_value_state(entry, source)
        src_phs = set(_PLACEHOLDER_RE.findall(src_value))
        if not src_phs:
            continue
        for lang in known:
            if lang == source:
                continue
            if entry_plural(entry, lang) is not None:
                continue  # 复数条目由 plural_problems 检查
            value, _ = entry_value_state(entry, lang)
            if not value:
                continue  # 缺译由 missing_problems 报
            phs = set(_PLACEHOLDER_RE.findall(value))
            if phs != src_phs:
                problems.append((lang, key,
                                 f"{path.stem}：占位符与源不一致"
                                 f"（源 {sorted(src_phs)} vs {sorted(phs)}）"))
    return problems


# --------------------------------------------------------------------------- #
# export 子命令
# --------------------------------------------------------------------------- #

def cmd_export(args: argparse.Namespace) -> int:
    if Workbook is None:
        sys.exit("错误：缺少 openpyxl，请执行 pip install -r tools/requirements.txt")

    explicit = [x for x in (args.lang.split(",") if args.lang else []) if x]
    out = Path(args.out)
    wb = Workbook()
    wb.remove(wb.active)  # 移除默认空 sheet

    for xc_raw in args.xcstrings:
        path = Path(xc_raw)
        obj = load_xcstrings(path)
        source = obj.get("sourceLanguage", "")

        # 确定导出语言列：显式指定优先，否则导出文件中已有语言
        if explicit:
            langs = [source] if source else []
            langs += [x for x in explicit if x != source]
        else:
            langs = collect_languages(obj)

        ws = wb.create_sheet(title=path.stem)
        header = ["key", "variation", "comment", "source_" + source if source else "source"]
        for lang in langs:
            if lang == source:
                continue
            header.append(lang)
            header.append(lang + "_state")
        ws.append(header)

        strings = obj.get("strings", {})
        var_order = {v: i for i, v in enumerate(PLURAL_VARIATION_ORDER)}
        for key in sorted(strings.keys()):
            entry = strings[key]
            comment = entry.get("comment", "")
            # 变体并集（含源语言）：任一语言为复数条目时按变体拆行；
            # 复数条目补齐语言要求变体（如 en/de/fr 的 one），即使当前尚无译文
            variants: set[str] = set()
            for lang in langs:
                rows_lang = entry_lang_rows(entry, lang)
                variants.update(var for var, _, _ in rows_lang if var)
            if variants:
                for lang in langs:
                    variants.update(plural_required_variations(lang))
            if not variants:
                row = [key, "", comment, entry_lang_rows(entry, source)[0][1]]
                for lang in langs:
                    if lang == source:
                        continue
                    value, state = entry_value_state(entry, lang)
                    row.append(value)
                    row.append(state)
                ws.append(row)
                continue
            by_src = {v: val for v, val, _ in entry_lang_rows(entry, source)}
            for var in sorted(variants, key=lambda v: var_order.get(v, 99)):
                row = [key, var, comment, by_src.get(var, "")]
                for lang in langs:
                    if lang == source:
                        continue
                    by_lang = {v: (val, st) for v, val, st in entry_lang_rows(entry, lang)}
                    val, st = by_lang.get(var, ("", ""))
                    row.append(val)
                    row.append(st)
                ws.append(row)

    if not wb.sheetnames:  # 兜底：无输入文件时不留空工作簿
        wb.create_sheet("empty")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"已导出 {len(args.xcstrings)} 个 String Catalog -> {out}")
    return 0


# --------------------------------------------------------------------------- #
# import 子命令
# --------------------------------------------------------------------------- #

def cmd_import(args: argparse.Namespace) -> int:
    if load_workbook is None:
        sys.exit("错误：缺少 openpyxl，请执行 pip install -r tools/requirements.txt")

    in_path = Path(args.in_file)
    xc_path = Path(args.xcstrings)
    obj = load_xcstrings(xc_path)
    source = obj.get("sourceLanguage", "")

    wb = load_workbook(in_path)
    if xc_path.stem not in wb.sheetnames:
        sys.exit(f"错误：Excel 中找不到 sheet「{xc_path.stem}」")
    rows = list(wb[xc_path.stem].iter_rows(values_only=True))
    if not rows:
        sys.exit("错误：sheet 为空")

    header = [str(h) if h is not None else "" for h in rows[0]]

    # 确定目标语言
    lang = args.lang
    if not lang:
        for h in header:
            if h and not h.startswith("source_") and not h.endswith("_state") and h not in ("key", "comment"):
                lang = h
                break
    if not lang:
        sys.exit("错误：未指定 --lang，且 Excel 中无目标语言列")
    if lang == source:
        sys.exit(f"错误：--lang（{lang}）与源语言相同，不可导入源语言列")

    val_idx = header.index(lang) if lang in header else None
    state_col = lang + "_state"
    state_idx = header.index(state_col) if state_col in header else None
    key_idx = header.index("key") if "key" in header else 0
    var_idx = header.index("variation") if "variation" in header else None
    if val_idx is None:
        sys.exit(f"错误：表头中找不到语言列「{lang}」")

    strings = obj.setdefault("strings", {})
    updated = added = 0
    for row in rows[1:]:
        if not row or row[key_idx] in (None, ""):
            continue
        key = str(row[key_idx]).strip()
        raw_var = row[var_idx] if var_idx is not None and len(row) > var_idx else None
        var = str(raw_var).strip() if raw_var else ""
        raw = row[val_idx]
        value = str(raw).strip() if raw is not None else ""
        if not value:  # 空单元格跳过，不覆盖已有译文
            continue
        raw_state = row[state_idx] if state_idx is not None else None
        state = str(raw_state).strip() if raw_state else "translated"

        entry = strings.get(key)
        if entry is None:
            entry = {}
            strings[key] = entry
            added += 1
        else:
            updated += 1
        loc_node = entry.setdefault("localizations", {}).setdefault(lang, {})
        if var:  # 复数变体行：与 stringUnit 互斥，合并进 variations.plural
            loc_node.pop("stringUnit", None)
            plural = loc_node.setdefault("variations", {}).setdefault("plural", {})
            plural["pluralRuleType"] = "pluralRuleType"
            plural[var] = {"stringUnit": {"state": state, "value": value}}
        else:  # 普通字符串行：替换为 stringUnit
            loc_node.pop("variations", None)
            loc_node["stringUnit"] = {"state": state, "value": value}

    save_xcstrings(xc_path, obj)
    print(f"已导入语言「{lang}」：更新 {updated} 条，新增 {added} 条 -> {xc_path}")
    return 0


# --------------------------------------------------------------------------- #
# check 子命令
# --------------------------------------------------------------------------- #

# 匹配 String(localized: "...") 与 NSLocalizedString("...")
_KEY_RE = re.compile(r'(?:String\(localized:|NSLocalizedString\()\s*"((?:\\.|[^"\\])*)"')
# 任意字符串字面量（Text("...") 等非 String(localized:) 引用）
_LITERAL_RE = re.compile(r'"((?:\\.|[^"\\])*)"')


def _replace_interpolations(key: str) -> str:
    """把 Swift 插值（\\(...) 平衡括号闭合）替换为 %lld。

    插值内部允许任意嵌套括号与三元表达式（如
    \\(pet.birthday != nil ? PetDisplayLogic.ageText(from: pet.birthday) : "—")），
    用深度计数找平衡右括号，不能用正则表达。
    """
    out: list[str] = []
    i = 0
    n = len(key)
    while i < n:
        if key[i] == "\\" and i + 1 < n and key[i + 1] == "(":
            depth = 1
            j = i + 2
            while j < n and depth > 0:
                if key[j] == "(":
                    depth += 1
                elif key[j] == ")":
                    depth -= 1
                j += 1
            out.append("%lld")
            i = j
        else:
            out.append(key[i])
            i += 1
    return "".join(out)


def normalize_localized_key(key: str) -> str:
    r"""归一化 key 用于代码↔目录比对：插值 \(x) 与格式占位符（%lld/%d/%@ 等）等价。

    手动复数 key 在目录中带 %lld 后缀（如 "paywall.cta.trial %lld"），
    代码侧以插值调用（"paywall.cta.trial \(days)"）；归一化后两者一致。
    """
    key = _replace_interpolations(key)
    key = re.sub(r"%(?:\d+\$)?[-+0-9.]*[a-zA-Z@]+", "%lld", key)
    return key


def extract_code_keys(source_root: Path) -> set[str]:
    """扫描 Swift 源码中引用的本地化 key。"""
    keys: set[str] = set()
    for swift in source_root.rglob("*.swift"):
        try:
            text = swift.read_text(encoding="utf-8")
        except (OSError, UnicodeDecodeError):
            continue
        keys.update(_KEY_RE.findall(text))
    return keys


def extract_literal_texts(source_root: Path) -> set[str]:
    """扫描 Swift 源码中的字符串字面量（Text/Button 直接文案等）。

    这些文案未走 String(localized:) API，但对应 String Catalog 的 manual key
    （key 以中文文案而非键名形式存在）会被字面量直接引用——check 需要把它们
    视为已引用，避免把仍在使用的文案误报为多余 key。
    """
    texts: set[str] = set()
    for swift in source_root.rglob("*.swift"):
        try:
            text = swift.read_text(encoding="utf-8")
        except (OSError, UnicodeDecodeError):
            continue
        texts.update(_LITERAL_RE.findall(text))
    return texts


# SwiftUI 中承载用户可见文案的 API（首参为字符串字面量）。
# Text("…") / Label("…",…) / Button("…") / navigationTitle("…") / .title("…") /
# TextField("…",…) / .help("…") / accessibilityLabel("…") / .prompt("…") 等。
# Section(header: Text("…")) 由 Text 分支覆盖。
_UI_LITERAL_RE = re.compile(
    r"\b(?:Text|Label|Button|NavigationLink|Link|MenuItem|Menu|"
    r"navigationTitle|toolbarTitle|title|help|accessibilityLabel|"
    r"accessibilityHint|accessibilityValue|prompt|tooltip)"
    r"\s*\(\s*\"((?:\\.|[^\"\\])*)\""
)
# CJK 统一表意范围（汉字 + 假名 + 韩文音节 + CJK 标点）：判定字面量是否含
# "本应本地化"的表意文字。纯英文/数字/符号串多为技术性标识（bundle id、
# 文件名、format 串、UserDefaults key），不含 CJK 故天然排除，不报。
_CJK_RE = re.compile(
    r"[\u3000-\u303f\u3040-\u309f\u30a0-\u30ff\u3400-\u4dbf"
    r"\u4e00-\u9fff\uac00-\ud7af\uf900-\ufaff]"
)


def _decode_swift_string(literal: str) -> str:
    """把 Swift 字符串字面量里的转义还原为实际字符（用于 catalog 比对）。"""
    return (literal.replace("\\n", "\n")
            .replace("\\t", "\t")
            .replace('\\"', '"')
            .replace("\\\\", "\\"))


def hardcoded_problems(source_root: Path, catalog_keys: set[str]
                       ) -> list[tuple[str, int, str]]:
    """扫描 Swift 源码中"疑似硬编码的用户可见文案" [(file, line, text)]。

    判定：出现在 SwiftUI 文案 API（Text/Label/Button/navigationTitle 等）
    首参位置的字符串字面量，含 CJK 表意文字，且不在 String Catalog（精确或
    插值/占位符归一化后均比对）。这类文案绕过了 String(localized:) 本地化
    管线，是日期/年龄/性别等动态内容被写死的高发区——翻译阶段无法覆盖。

    误报控制：仅匹配已知 UI API 上下文（非任意字面量）+ 必须含 CJK 字符。
    结果为"疑似"，需人工复核后收口进 catalog 或改 String(localized:)。
    """
    norm_keys = {normalize_localized_key(k) for k in catalog_keys}
    problems: list[tuple[str, int, str]] = []
    for swift in source_root.rglob("*.swift"):
        try:
            text = swift.read_text(encoding="utf-8")
        except (OSError, UnicodeDecodeError):
            continue
        for m in _UI_LITERAL_RE.finditer(text):
            literal = m.group(1)
            if not _CJK_RE.search(literal):
                continue
            decoded = _decode_swift_string(literal)
            if decoded in catalog_keys:
                continue
            if normalize_localized_key(decoded) in norm_keys:
                continue
            line = text.count("\n", 0, m.start()) + 1
            problems.append((str(swift.relative_to(source_root)), line, decoded))
    return problems


def parse_known_regions(project_yml: Path) -> list[str]:
    """从 project.yml 提取 knownRegions（简单正则，避免 pyyaml 依赖）。"""
    if not project_yml.exists():
        return []
    text = project_yml.read_text(encoding="utf-8")
    m = re.search(r"knownRegions:\s*\[([^\]]*)\]", text)
    if not m:
        return []
    return [r.strip().strip("\"'") for r in m.group(1).split(",") if r.strip()]


_LENGTH_TAG_RE = re.compile(r"\[len:\s*(\d+)\]")


def load_length_rules(path: Path | None) -> dict:
    """加载长度规则 JSON。无文件返回空规则（不检查）。

    格式：{"default": int, "keys": {key: int}, "prefixes": {prefix: int},
    "comment_tag": "len"}。规则优先级：精确 key > comment [len:N] 标记 >
    最长匹配前缀 > default。
    """
    if not path:
        return {}
    try:
        with path.open(encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        raise SystemExit(f"错误：长度规则文件 {path} 解析失败 -> {e}")
    return data if isinstance(data, dict) else {}


def resolve_maxlen(key: str, comment: str, rules: dict) -> int | None:
    """解析某 key 的最大长度上限；无规则返回 None（不检查）。"""
    keys = rules.get("keys", {})
    if key in keys:
        return keys[key]
    tag = rules.get("comment_tag", "len")
    if tag and comment:
        m = _LENGTH_TAG_RE.search(comment)
        if m:
            return int(m.group(1))
    prefixes = rules.get("prefixes", {})
    best: int | None = None
    best_len = -1
    for prefix, mx in prefixes.items():
        if key.startswith(prefix) and len(prefix) > best_len:
            best, best_len = mx, len(prefix)
    if best is not None:
        return best
    d = rules.get("default")
    return int(d) if d else None


def length_problems(obj: dict, path: Path, known: list[str], rules: dict
                    ) -> list[tuple[str, str, str]]:
    """译文长度超限清单 [(lang, key, note)]；rules 为空时直接返回 []。

    复数条目取每个变体长度判定（德语等长语言各变体都可能超限）。
    """
    if not rules:
        return []
    problems: list[tuple[str, str, str]] = []
    for key in sorted(obj.get("strings", {}).keys()):
        entry = obj["strings"][key]
        mx = resolve_maxlen(key, entry.get("comment", ""), rules)
        if not mx:
            continue
        for lang in known:
            if lang == obj.get("sourceLanguage", ""):
                continue
            for var, value, _state in entry_lang_rows(entry, lang):
                if not value:
                    continue
                if len(value) > mx:
                    tag = f"「{var}」" if var else ""
                    problems.append((lang, key,
                                     f"{path.stem}：{tag}长度 {len(value)} 超过 {mx}"))
    return problems


def cmd_check(args: argparse.Namespace) -> int:
    # 新参数用 getattr 兜底，保证 GUI run_check（只传基础参数）兼容
    strict = getattr(args, "strict", False)
    length_rules_path = getattr(args, "length_rules", None)
    do_hardcoded = getattr(args, "hardcoded", False)
    rules = load_length_rules(Path(length_rules_path) if length_rules_path else None)

    problems = 0     # 阻断级问题数（决定退出码）
    warnings = 0     # 警告级问题数（多余 key / needs_review 非 strict，不阻断）
    project_yml = Path(args.project_yml) if args.project_yml else None
    source_root = Path(args.source_root) if args.source_root else None
    code_keys = extract_code_keys(source_root) if source_root else set()
    code_literals = extract_literal_texts(source_root) if source_root else set()
    known = parse_known_regions(project_yml) if project_yml else []
    loaded: list[tuple[Path, dict]] = []
    last_source = ""

    for xc_raw in args.xcstrings:
        path = Path(xc_raw)
        # ① JSON 合法性
        try:
            obj = load_xcstrings(path)
        except json.JSONDecodeError as e:
            print(f"[失败] {path}: JSON 解析错误 -> {e}")
            problems += 1
            continue
        except OSError as e:
            print(f"[失败] {path}: 无法读取 -> {e}")
            problems += 1
            continue
        loaded.append((path, obj))

        source = obj.get("sourceLanguage", "")
        last_source = source
        strings = obj.get("strings", {})
        present = collect_languages(obj)
        print(f"[OK] {path.name}: {len(strings)} key，语言 {present}")

        # ② 格式稳定性（load -> dump 应与磁盘内容一致）
        original = path.read_text(encoding="utf-8")
        if original != dumps_xcstrings(obj):
            print(f"  [格式] {path.name} 与规范化（Xcode 风格）不一致，运行 import 或 Xcode 保存可修正")

        # ③ 代码引用一致性（仅 Localizable.xcstrings）
        if source_root and path.stem == "Localizable":
            norm_literals = {normalize_localized_key(t) for t in code_literals}
            norm_xc = {normalize_localized_key(k) for k in strings.keys()}
            missing = sorted(k for k in code_keys if normalize_localized_key(k) not in norm_xc)
            # 多余 key 排除字面量引用：key 以中文文案形式存在时，代码常以
            # Text("...") 直接字面量引用（未走 String(localized:) API）；
            # 比较前把 JSON 解码的 \n 还原为源码 \n 文本、\" 还原为 \\",
            # 并对插值/占位符做归一化（复数 key 的 %lld 与 \\(days) 等价）。
            extra = set()
            for key in strings.keys():
                if key in code_keys:
                    continue
                source_form = key.replace("\n", "\\n").replace('"', '\\"')
                if source_form not in code_literals and normalize_localized_key(source_form) not in norm_literals:
                    extra.add(key)
            if missing:
                print(f"  [缺 key] 代码引用但 String Catalog 缺少：{missing}")
                problems += len(missing)
            if extra:
                print(f"  [多余 key] String Catalog 有但代码未引用：{sorted(extra)}")
                warnings += len(extra)

        # ④ 缺译检测 + 复数完整性（knownRegions 非源语言）——阻断
        for lang, key, note in missing_problems(obj, path, known):
            print(f"  [缺译] {key} @ {lang}: {note}")
            problems += 1

        # ⑤ 普通条目占位符漂移——阻断：译文漏 %@/%d 会运行时崩或显示畸形
        for lang, key, note in placeholder_problems(obj, path, known):
            print(f"  [占位符] {key} @ {lang}: {note}")
            problems += 1

        # ⑥ 初译待审（needs_review）—— --strict 阻断，否则警告
        for lang, key, note in review_problems(obj, path, known):
            print(f"  [待审] {key} @ {lang}: {note}")
            if strict:
                problems += 1
            else:
                warnings += 1

        # ⑧ 译文长度超限（--length-rules 开启）——阻断
        for lang, key, note in length_problems(obj, path, known, rules):
            print(f"  [超长] {key} @ {lang}: {note}")
            problems += 1

    # ⑦ 每语言进度统计（合并全部 xcstrings）
    if known and loaded:
        print("\n[进度] 各语言翻译统计：")
        for st in scan_statuses(loaded, known):
            if st.is_source:
                mark = "（源语言）"
            elif st.missing:
                mark = "缺译"
            elif st.needs_review:
                mark = "待审"
            else:
                mark = "完成"
            print(f"  {st.lang:8s} {st.name:8s} "
                  f"total={st.total:<4d} ok={st.translated:<4d} "
                  f"review={st.needs_review:<3d} missing={st.missing:<4d} "
                  f"({st.pct}%) {mark}")
        non_source = [r for r in known if r != last_source]
        if non_source:
            print(f"[信息] knownRegions={known}，源语言={last_source}；"
                  f"非源语言 {non_source} 需在各 xcstrings 补齐译文。")

    # ⑨ 硬编码用户可见文案检测（--hardcoded 开启）——阻断
    if do_hardcoded and source_root and loaded:
        all_keys: set[str] = set()
        for _p, o in loaded:
            all_keys.update(o.get("strings", {}).keys())
        hards = hardcoded_problems(source_root, all_keys)
        for f, line, text in hards:
            print(f"  [硬编码] {f}:{line} 疑似未本地化文案：{text!r}")
        problems += len(hards)

    if problems or warnings:
        parts = [f"{problems} 个阻断问题"] if problems else []
        if warnings:
            parts.append(f"{warnings} 个警告")
        tail = "（警告不阻断，加 --strict 升级 needs_review 为阻断）" if (warnings and not strict) else ""
        print(f"\n校验完成：{'，'.join(parts)} {tail}".rstrip())
        return 1 if problems else 0
    print("\n校验通过。")
    return 0


# --------------------------------------------------------------------------- #
# 入口
# --------------------------------------------------------------------------- #

def main() -> None:
    # Windows 默认控制台编码（GBK）无法输出 − 等 Unicode 字符，check 会抛
    # UnicodeEncodeError（此前需 PYTHONUTF8=1 才通过）；统一重配置 stdout 为 UTF-8，
    # 等价于设置 PYTHONUTF8=1，Windows/macOS/Linux 行为一致。
    if sys.stdout is not None:
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except (AttributeError, ValueError):
            # 非 TextIOWrapper（如测试环境替换的流）或不可重配置时保持原样
            pass

    p = argparse.ArgumentParser(
        description="MiLens 本地化（String Catalog）导出 / 导入 / 校验工具。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="详见 DEVELOPMENT.md §4.4。",
    )
    sub = p.add_subparsers(dest="command", required=True)

    pe = sub.add_parser("export", help="导出 xcstrings 为 Excel")
    pe.add_argument("xcstrings", nargs="+", help="一个或多个 .xcstrings 文件")
    pe.add_argument("out", help="输出 .xlsx 路径")
    pe.add_argument("--lang", help="要导出的目标语言，逗号分隔（如 en,ja）；不指定则导出文件中已有语言")
    pe.set_defaults(func=cmd_export)

    pi = sub.add_parser("import", help="从 Excel 导入回 xcstrings")
    pi.add_argument("in_file", help="输入 .xlsx 路径")
    pi.add_argument("xcstrings", help="目标 .xcstrings 文件")
    pi.add_argument("--lang", help="要导入的目标语言；不指定则取第一个非源语言列")
    pi.set_defaults(func=cmd_import)

    pc = sub.add_parser("check", help="校验 xcstrings 合法性与一致性")
    pc.add_argument("xcstrings", nargs="+", help="一个或多个 .xcstrings 文件")
    pc.add_argument("--project-yml", help="project.yml 路径（用于对比 knownRegions）")
    pc.add_argument("--source-root", help="Swift 源码根目录（扫描代码引用的 key）")
    pc.add_argument("--strict", action="store_true",
                    help="发布门禁：把 needs_review 初译待审也计为阻断问题")
    pc.add_argument("--length-rules", metavar="JSON",
                    help="译文长度规则 JSON 路径（{default,keys,prefixes,comment_tag}）")
    pc.add_argument("--hardcoded", action="store_true",
                    help="扫描 Swift 源码中疑似硬编码的用户可见文案（含 CJK 且不在 catalog）")
    pc.set_defaults(func=cmd_check)

    args = p.parse_args()
    sys.exit(args.func(args))


if __name__ == "__main__":
    main()
