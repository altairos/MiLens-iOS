#!/usr/bin/env python3
"""MiLens 全球本地化资产工作簿生成器。

产出 docs/localization/global-localization.xlsx，涵盖全部 6 类本地化资产 × 7 语种：

- Localizable / InfoPlist sheet：复用 tools/localization.py 的导出格式（列结构
  完全一致），翻译完成后可直接用 `localization.py import` 回写 .xcstrings；
  工作簿中其他 sheet 不影响 import（工具按 sheet 名匹配，只取这两个）。
- 其余 sheet（状态总览 / AppStore元数据 / 订阅产品描述 / 隐私政策 / 截图素材 /
  ASO关键词 / 术语表）：资产清单，人工在 Excel 中维护，录入 App Store Connect
  时人工搬运。

重新生成（幂等，已有译文保留在 .xcstrings 中，不会被覆盖）：
  python tools/localization-assets.py

依赖：openpyxl（pip install -r tools/requirements.txt）。
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")  # Windows GBK 控制台无需 PYTHONUTF8=1

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import localization as loc

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "docs" / "localization"
OUT = OUT_DIR / "global-localization.xlsx"
XCSTRINGS = [
    ROOT / "MiLens" / "Resources" / "Localizable.xcstrings",
    ROOT / "MiLens" / "Resources" / "InfoPlist.xcstrings",
]

# 首发 6 种非源语言（源语言 zh-Hans 固定第一列）
LANGS = ["zh-Hant", "ja", "ko", "en", "fr", "de"]

# --------------------------------------------------------------------------- #
# 样式
# --------------------------------------------------------------------------- #

HEADER_FILL = PatternFill("solid", fgColor="F3EAE2")
HEADER_FONT = Font(bold=True)
STATE_FONT = Font(color="8A7F75")
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"

def autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# --------------------------------------------------------------------------- #
# 1) Localizable / InfoPlist sheet —— 工具链原生格式（localization.py export 同构）
# --------------------------------------------------------------------------- #

def sheet_from_xcstrings(wb: Workbook, path: Path) -> None:
    obj = loc.load_xcstrings(path)
    source = obj.get("sourceLanguage", "")
    langs = [source] + [x for x in LANGS if x != source]

    ws = wb.create_sheet(title=path.stem)
    header = ["key", "comment", "source_" + source]
    for lang in langs:
        if lang == source:
            continue
        header.append(lang)
        header.append(lang + "_state")
    ws.append(header)

    for key in sorted(obj.get("strings", {}).keys()):
        entry = obj["strings"][key]
        comment = entry.get("comment", "")
        src_val, _ = loc.entry_value_state(entry, source)
        row = [key, comment, src_val or key]  # 字面量 key 无显式译文，源文案=key 本身
        for lang in langs:
            if lang == source:
                continue
            value, state = loc.entry_value_state(entry, lang)
            if lang == "en" and not value:
                # en 初译（批次 0 起点）：仅回填空值，已有译文以 .xcstrings 为准
                value = EN_TRANSLATIONS.get(key, "")
                state = "needs_review" if value else ""
            row.append(value)
            row.append(state)
        ws.append(row)

    style_header(ws, len(header))
    autosize(ws, [46, 34, 46] + [46, 14] * (len(langs) - 1))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column % 2 == 0 or cell.column == 1:
                cell.alignment = WRAP
            if cell.value is None:
                cell.value = ""

# --------------------------------------------------------------------------- #
# 2) 状态总览 sheet —— 6 类资产 × 7 语言
# --------------------------------------------------------------------------- #

OVERVIEW = [
    # (资产, 行数说明, {语言: 状态})
    ("1. App UI 文案（Localizable ~150 key）", "144 key，String Catalog", {
        "zh-Hans": "✅ 源语言", "zh-Hant": "待翻译", "ja": "待翻译", "ko": "待翻译",
        "en": "初译待审（needs_review）", "fr": "待翻译", "de": "待翻译"}),
    ("2. 权限说明 + App 显示名（InfoPlist 3 key）", "权限文案 + CFBundleDisplayName", {
        "zh-Hans": "✅ 源语言", "zh-Hant": "待翻译", "ja": "待翻译", "ko": "待翻译",
        "en": "初译待审（needs_review）", "fr": "待翻译", "de": "待翻译"}),
    ("3. App Store 元数据", "名称/副标题/描述/关键词/推广文本/审核备注", {
        "zh-Hans": "✅ 已定稿", "zh-Hant": "待制作", "ja": "待制作", "ko": "待制作",
        "en": "待制作", "fr": "待制作", "de": "待制作"}),
    ("4. 订阅产品描述（ASC）", "3 产品 × 显示名/描述", {
        "zh-Hans": "✅ 已定稿", "zh-Hant": "待制作", "ja": "待制作", "ko": "待制作",
        "en": "待制作", "fr": "待制作", "de": "待制作"}),
    ("5. 隐私政策网页", "docs/privacy-policy.html，?lang= 切换", {
        "zh-Hans": "✅ 已发布", "zh-Hant": "待制作", "ja": "待制作", "ko": "待制作",
        "en": "待制作", "fr": "待制作", "de": "待制作"}),
    ("6. 截图/预览素材", "3 设备尺寸 × 5 张 × 文案覆盖层", {
        "zh-Hans": "待制作", "zh-Hant": "待制作", "ja": "待制作", "ko": "待制作",
        "en": "待制作", "fr": "待制作", "de": "待制作"}),
]

def add_overview(wb: Workbook) -> None:
    ws = wb.create_sheet(title="状态总览")
    ws.append(["资产", "范围", *LANGS])
    for asset, scope, states in OVERVIEW:
        ws.append([asset, scope] + [states.get(l, "") for l in LANGS])
    style_header(ws, 3 + len(LANGS))
    autosize(ws, [40, 40] + [22] * len(LANGS))

# --------------------------------------------------------------------------- #
# 3) App Store 元数据 sheet —— 6 字段 × 7 语言
# --------------------------------------------------------------------------- #

# 每字段：中文定稿（AppStore-metadata.md），其余语言草案/空
APP_NAME = {
    "zh-Hans": "咪Lens - 宠物照片整理与拼豆创作",
    "zh-Hant": "咪Lens - 寵物照片整理與拼豆創作",
    "ja": "咪Lens - ペット写真整理とアイロンビーズ図案",
    "ko": "MiLens - 반려동물 사진 정리와 비즈 도안",
    "en": "MiLens - Pet Photos & Bead Art",
    "fr": "MiLens - Photos animaux & perles à repasser",
    "de": "MiLens - Haustierfotos & Bügelperlen-Vorlagen",
}
SUBTITLE = {
    "zh-Hans": "拾回散落的每一张照片，记住你与爱宠共度的一生",
    "zh-Hant": "拾回散落的每一張照片，記住你與毛小孩共度的一生",
    "ja": "ペットの写真を整理して、アイロンビーズの図案に",
    "ko": "반려동물 사진 정리와 비즈 도안 만들기",
    "en": "Every pet photo, organized & turned into bead art",
    "fr": "Triez les photos d'animaux, créez des perles à repasser",
    "de": "Haustierfotos sortieren & Bügelperlen-Vorlagen",
}
DESCRIPTION_ZH = """咪Lens 是宠物家庭的数字生命档案。

它会帮你找到手机里的宠物照片，为每只宠物建立档案，保存它们的成长回忆，还能把照片变成拼豆图纸——让这份陪伴以另一种形式留在身边。

【自动发现】
扫描相册，AI 自动找出宠物照片；你可以把照片归到每只宠物的档案里。照片不会离开手机，所有分析都在本地完成。

【成长回忆】
记录每只宠物的档案，自动生成成长时间线。首次打开后的 14 天内可完整浏览历史，之后免费版查看最近一年；MiLens Pro 帮你找回更早的故事。

【照片整理】
质量评分帮你筛选最清晰的照片，重复检测自动标记连拍和相似图。不再在几百张照片里翻找。

【拼豆创作】
把宠物照片变成拼豆图纸。免费版每天可生成 5 张，MiLens Pro 不限次数。智能色彩匹配，多种风格与配色方案可选。

【完整编辑器】
裁切、调色、抠图与文字工具免费使用，让每张照片都更接近你记忆里的样子。

免费版可创建 1 个宠物档案、保存 50 张照片；MiLens Pro 支持最多 20 个档案、不限照片存储、不限拼豆生成和完整成长历史，导出作品无水印。

【隐私至上】
所有照片分析在设备本地完成。App 不会主动上传照片；编辑产物可能随用户启用的系统备份保存。

---

MiLens Pro 订阅：
- 月度订阅：每月自动续费
- 年度订阅：每年自动续费，约合每月 8 元
- 永久版：一次购买，永久解锁当前 Pro 权益

V1.0 计划加入高级创作模板与高清导出，Pro 用户上线后自动解锁。

订阅会自动续费，除非在当前周期结束前至少 24 小时关闭自动续费。
管理订阅或取消可在 App Store 账户设置中操作。

隐私政策：[待填]
服务条款：[待填]"""
KEYWORDS = {
    "zh-Hans": "宠物照片,宠物相册,猫,狗,照片整理,拼豆,拼豆图纸,照片管理,宠物档案,成长记录",
    "zh-Hant": "寵物照片,寵物相簿,拼豆,拼豆圖紙,毛小孩,貓咪,狗狗,照片整理",
    "ja": "ペット写真,アイロンビーズ,パーラービーズ,写真整理,猫,犬,ペットアルバム,ペット写真整理",
    "ko": "반려동물,강아지,고양이,비즈 도안,사진 정리,반려동물 앨범",
    "en": "pet photos,pet album,perler beads,bead pattern,photo organizer,cat,dog",
    "fr": "perles à repasser,photos animaux,album chat,album chien",
    "de": "Bügelperlen,Bügelperlen Vorlagen,Haustier,Hund,Katze,Fotoalbum,Fotos sortieren",
}
PROMO_ZH = "用 AI 找出手机里每一张宠物照片，自动整理成专属相册。还能把照片变成拼豆图纸！免费试用 7 天 MiLens Pro。"
REVIEW_ZH = """咪Lens 是一款宠物照片管理和创作 App。

审核说明：
1. 扫描功能需要访问照片库权限——请允许后等待扫描完成。
2. AI 宠物识别在设备本地完成（Core ML + Vision），不需要网络连接。
3. 首次使用需要创建至少一个宠物档案才能看到照片分类效果。
4. 拼豆图纸生成需要从相册选择一张照片后进入创作 Tab。
5. 免费版可创建 1 个宠物档案、每天生成 5 张拼豆图纸；首次打开时间线后 14 天内可查看完整历史，之后显示最近一年。MiLens Pro 解锁 20 个档案、不限生成和完整历史。图片编辑器当前免费。

测试账号：无需登录，App 不含任何账号系统。"""

META_FIELDS = [
    ("名称（30 字符）", APP_NAME),
    ("副标题（30 字符）", SUBTITLE),
    ("描述（4000 字符）", {k: (DESCRIPTION_ZH if k == "zh-Hans" else "") for k in ["zh-Hans"] + LANGS}),
    ("关键词（100 字符）", KEYWORDS),
    ("推广文本（170 字符）", {k: (PROMO_ZH if k == "zh-Hans" else "") for k in ["zh-Hans"] + LANGS}),
    ("审核备注", {k: (REVIEW_ZH if k == "zh-Hans" else "") for k in ["zh-Hans"] + LANGS}),
]

def add_metadata(wb: Workbook) -> None:
    ws = wb.create_sheet(title="AppStore元数据")
    ws.append(["字段", *["zh-Hans"] + LANGS])
    for name, per_lang in META_FIELDS:
        ws.append([name] + [per_lang.get(l, "") for l in ["zh-Hans"] + LANGS])
    style_header(ws, 1 + 7)
    autosize(ws, [30] + [60] * 7)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column >= 2:
                cell.alignment = WRAP

# --------------------------------------------------------------------------- #
# 4) 订阅产品描述 sheet —— 3 产品 × (显示名/描述) × 7 语言
# --------------------------------------------------------------------------- #

SUBS = [
    ("com.milens.pro.monthly", "月度订阅", "MiLens Pro 月度订阅", """解锁当前 Pro 权益：
• 最多 20 个宠物档案
• 拼豆图纸不限次数生成
• 查看完整成长时间线
• 支持 Apple 家庭共享

V1.0 计划加入高级创作模板与高清导出，Pro 用户上线后自动解锁。

每月自动续费，可随时取消。"""),
    ("com.milens.pro.yearly", "年度订阅", "MiLens Pro 年度订阅", """解锁当前 Pro 权益：
• 最多 20 个宠物档案
• 拼豆图纸不限次数生成
• 查看完整成长时间线
• 支持 Apple 家庭共享

V1.0 计划加入高级创作模板与高清导出，Pro 用户上线后自动解锁。

每年自动续费，约合每月 8 元，比月度节省 55%。可随时取消。"""),
    ("com.milens.pro.lifetime", "永久版", "MiLens Pro 永久版", """一次购买，永久解锁当前 Pro 权益：
• 最多 20 个宠物档案
• 拼豆图纸不限次数生成
• 查看完整成长时间线
• 支持 Apple 家庭共享

V1.0 计划加入高级创作模板与高清导出，Pro 用户上线后自动解锁。

一次付费，无后续费用，永久使用。"""),
]

def add_subscriptions(wb: Workbook) -> None:
    ws = wb.create_sheet(title="订阅产品描述")
    ws.append(["产品 ID", "产品", "字段", *["zh-Hans"] + LANGS])
    for pid, name, display, desc in SUBS:
        ws.append([pid, name, "显示名", display, "", "", "", "", "", ""])
        ws.append([pid, name, "描述（ASC 本地化）", desc, "", "", "", "", "", ""])
    style_header(ws, 3 + 7)
    autosize(ws, [26, 12, 22] + [60] * 7)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column >= 4:
                cell.alignment = WRAP

# --------------------------------------------------------------------------- #
# 5) 隐私政策 sheet —— 章节 × 7 语言
# --------------------------------------------------------------------------- #

PRIVACY_SECTIONS = [
    ("标题", "咪Lens 隐私政策"),
    ("主体", "昆明妙谧智能科技有限责任公司"),
    ("日期", "最后更新：2026年8月8日"),
    ("一、我们收集什么", "咪Lens 不要求注册账号，不收集姓名、邮箱、手机号等身份信息。"),
    ("二、我们不做什么", "不主动上传照片——照片在设备本地读取、分析、编辑，App 不会将照片发送到网络服务器；用户启用的系统备份可能保存编辑产物。"),
    ("三、AI 分析如何工作", "咪Lens 使用设备本地 AI 技术分析宠物照片。"),
    ("四、第三方框架", "咪Lens 仅使用 Apple 官方系统框架，不包含任何第三方分析 SDK、广告 SDK 或追踪 SDK。"),
    ("五、数据存储", "所有 App 数据存储在你的设备本地。"),
    ("六、儿童隐私", "咪Lens 不针对 13 岁以下儿童，也不收集儿童信息。内容评级 4+。"),
    ("七、隐私政策变更", "如果我们更新隐私政策，会在此页面发布最新版本并更新日期。"),
    ("联系我们", "昆明妙谧智能科技有限责任公司 · privacy@miovelle.cn"),
]

def add_privacy(wb: Workbook) -> None:
    ws = wb.create_sheet(title="隐私政策")
    ws.append(["章节（privacy-policy.html）", *["zh-Hans"] + LANGS])
    for section, zh in PRIVACY_SECTIONS:
        ws.append([section, zh, "", "", "", "", "", ""])
    style_header(ws, 1 + 7)
    autosize(ws, [34] + [60] * 7)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column >= 2:
                cell.alignment = WRAP

# --------------------------------------------------------------------------- #
# 6) 截图素材 sheet —— 5 张截图 × (画面/文案) × 7 语言
# --------------------------------------------------------------------------- #

SCREENSHOTS = [
    ("1 扫描发现页", "找到手机里每一张宠物照片"),
    ("2 宠物档案页", "认识你的每一只宠物"),
    ("3 相册网格页", "自动整理，质量评分"),
    ("4 拼豆图纸生成", "把照片变成拼豆作品"),
    ("5 成长时间线", "保存它的一生"),
]

def add_screenshots(wb: Workbook) -> None:
    ws = wb.create_sheet(title="截图素材")
    ws.append(["截图", "字段", *["zh-Hans"] + LANGS])
    for shot, zh_text in SCREENSHOTS:
        ws.append([shot, "画面说明", shot.split(" ", 1)[1]])
        ws.append([shot, "叠加文案", zh_text, "", "", "", "", "", ""])
    style_header(ws, 2 + 7)
    autosize(ws, [20, 14] + [40] * 7)

# --------------------------------------------------------------------------- #
# 7) ASO 关键词 sheet —— 每语言关键词清单
# --------------------------------------------------------------------------- #

ASO_ROWS = [
    ("zh-Hans", KEYWORDS["zh-Hans"], "拼豆图纸（低竞争高意图）", "中"),
    ("zh-Hant", KEYWORDS["zh-Hant"], "拼豆圖紙、寵物相簿", "低"),
    ("ja", KEYWORDS["ja"], "アイロンビーズ図案（ビーズ词族竞争低）", "中（ビーズ低）"),
    ("ko", KEYWORDS["ko"], "비즈 도안、반려동물 앨범", "中"),
    ("en", KEYWORDS["en"], "perler bead pattern（长尾）", "高"),
    ("fr", KEYWORDS["fr"], "perles à repasser modèle", "低"),
    ("de", KEYWORDS["de"], "Bügelperlen Vorlagen（强意图）", "低-中"),
]

def add_aso(wb: Workbook) -> None:
    ws = wb.create_sheet(title="ASO关键词")
    ws.append(["语言", "关键词（100 字符内，逗号分隔）", "长尾机会", "竞争度"])
    for lang, words, tail, comp in ASO_ROWS:
        ws.append([lang, words, tail, comp])
    style_header(ws, 4)
    autosize(ws, [12, 80, 40, 16])

# --------------------------------------------------------------------------- #
# 8) 术语表 sheet —— 核心术语 × 7 语言（Localization-Plan.md §8）
# --------------------------------------------------------------------------- #

GLOSSARY = [
    ("拼豆图纸", "bead pattern", "アイロンビーズ図案", "비즈 도안", "Bügelperlen-Vorlage", "modèle de perles à repasser", "拼豆圖紙"),
    ("拼豆工作室", "bead studio", "ビーズスタジオ", "비즈 스튜디오", "Bead-Studio", "studio de perles", "拼豆工作室"),
    ("宠物档案", "pet profile", "ペットのプロフィール", "반려동물 프로필", "Haustierprofil", "profil de l'animal", "寵物檔案"),
    ("成长时间线", "growth timeline", "成長タイムライン", "성장 타임라인", "Wachstumszeitleiste", "chronologie de croissance", "成長時間軸"),
    ("往日回忆", "memories", "思い出", "추억", "Erinnerungen", "souvenirs", "往日回憶"),
    ("时光机", "time machine", "タイムマシン", "타임머신", "Zeitmaschine", "machine à remonter le temps", "時光機"),
    ("照片不离开设备", "photos stay on device", "写真は端末から出ません", "모든 분석은 기기에서 처리됩니다", "Alle Fotos bleiben auf Ihrem Gerät", "traitement 100 % local", "照片不會離開手機"),
]

def add_glossary(wb: Workbook) -> None:
    ws = wb.create_sheet(title="术语表")
    ws.append(["中文（源）", "en", "ja", "ko", "de", "fr", "zh-Hant"])
    for row in GLOSSARY:
        ws.append(list(row))
    style_header(ws, 7)
    autosize(ws, [24, 32, 34, 34, 36, 36, 26])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP

# --------------------------------------------------------------------------- #
# en 初译（批次 0 起点，state=needs_review，翻译人员可在 Excel 中直接修改后 import）
# --------------------------------------------------------------------------- #

EN_TRANSLATIONS = {
    # 显式 key（带 localizations）
    "about.fonts.note": "All fonts are licensed under the SIL Open Font License 1.1; the license text is distributed with the app.",
    "about.fonts.title": "Fonts & Open Source Licenses",
    "help.a1": "Tap \u201cScan\u201d on the album page \u2014 MiLens finds photos that may contain pets on your device, then you confirm and import them.",
    "help.a2": "Pick a pet when confirming scan results; you can also select multiple photos in the album and file them under a pet.",
    "help.a3": "Go to the Create tab, choose \u201cBead Pattern\u201d, pick a photo, adjust size and style, then generate and export.",
    "help.a4": "Bead patterns can be saved to your photo library or shared with others via the system share sheet.",
    "help.q1": "How do I find pet photos on my phone?",
    "help.q2": "How do I file photos under a pet profile?",
    "help.q3": "How do I make a bead pattern?",
    "help.q4": "Where are my exported works saved?",
    "paywall.alert.ok": "OK",
    "paywall.alert.title": "Notice",
    "paywall.badge.recommended": "Recommended",
    "paywall.benefit.beadQuota": "Unlimited bead pattern generation",
    "paywall.benefit.family": "Family Sharing supported (via Apple StoreKit)",
    "paywall.benefit.profiles": "Manage up to 20 pet profiles",
    "paywall.benefit.timeline": "Full growth timeline, including stories from years ago",
    "paywall.benefit.photoStorage": "Unlimited pet photo storage",
    "paywall.benefit.watermarkFree": "Watermark-free exports",
    "paywall.benefit.cardTemplates": "Beautiful card templates",
    "paywall.benefit.timelineExport": "Export & share your growth timeline",
    "paywall.benefit.offlineBackup": "Offline photo backup & restore",
    "paywall.benefit.albumModes": "Vintage album browsing modes",
    "paywall.close": "Close",
    "paywall.cta.lifetime.price": "Unlock Forever \u00b7 %@",
    "paywall.cta.subscribe.price": "Subscribe Now \u00b7 %@",
    "paywall.cta.trial": "Free trial \u00b7 %d days",
    "paywall.cta.unavailable": "Unavailable",
    "paywall.future.body": "Advanced templates and HD export are planned for V1.0 and will unlock automatically for Pro users.",
    "paywall.link.privacy": "Privacy Policy",
    "paywall.link.terms": "Terms of Service",
    "paywall.load.failed": "Unable to load subscription info. Check your connection and try again.",
    "paywall.period.monthly": "Monthly",
    "paywall.period.once": "One-time",
    "paywall.period.yearly": "Yearly",
    "paywall.pro.owned": "You've unlocked MiLens Pro",
    "paywall.purchase.failed": "Purchase didn't complete. Please try again later.",
    "paywall.purchase.pending": "Purchase is pending confirmation and will unlock automatically once complete.",
    "paywall.restore": "Restore Purchase",
    "paywall.restore.done": "Your purchase has been restored.",
    "paywall.restore.failed": "Restore failed. Please try again later.",
    "paywall.restore.none": "No purchases found to restore.",
    "paywall.retry": "Retry",
    "paywall.subtitle": "Unlock every creative feature in one go",
    "paywall.terms.lifetime": "Pay once, use forever. No recurring fees.",
    "paywall.terms.subscription": "Renews automatically at %@. Cancel anytime in your App Store account settings; your current period remains valid after cancellation.",
    "paywall.terms.trial": "Free for %1$d days, then renews at %2$@. Cancel anytime in your App Store account settings; your current period remains valid after cancellation.",
    "paywall.title": "Turn this photo into a keepsake",
    "paywall.trial.hint": "First %d days free",
    "privacy.commit.control": "You decide what gets imported",
    "privacy.commit.control.detail": "Scanning only finds candidate photos; you confirm which ones to import and which pet they belong to.",
    "privacy.commit.local": "Analysis happens on your device",
    "privacy.commit.local.detail": "Pet photo recognition runs on on-device models \u2014 no servers involved.",
    "privacy.commit.ondevice": "Photos are never uploaded",
    "privacy.commit.ondevice.detail": "The app never uploads your photos; edited works may be saved by system backups you've enabled.",
    "privacy.commit.title": "On-device privacy promise",
    "privacy.permissions.manage": "Manage photo & notification permissions",
    "privacy.policy.link": "View Privacy Policy",
    "privacy.storage.detail": "Imported photo copies and profile data are stored in the app's sandbox. Deleting the app removes this data; originals in your system photo library are unaffected.",
    "privacy.storage.title": "Data Storage",
    "settings.about.entry": "About MiLens",
    "settings.about.version": "Version",
    "settings.appearance.dark": "Dark",
    "settings.appearance.light": "Light",
    "settings.appearance.system": "System",
    "settings.appearance.title": "Appearance",
    "settings.notifications.denied.message": "Allow notifications for MiLens in System Settings, then enable memorial reminders.",
    "settings.notifications.denied.ok": "OK",
    "settings.notifications.denied.title": "Notifications Disabled",
    "settings.notifications.footer": "When enabled: reminders at 9:00 AM on your pet's birthday and adoption day; a daily 9:00 AM time-machine memory when photos exist from this day in past years.",
    "settings.notifications.reminders": "Memorial Reminders",
    "settings.privacy.local": "Local data & privacy",
    "settings.privacy.policy": "Privacy Policy",
    "settings.pro.active.title": "MiLens Pro Active",
    "settings.pro.manage": "Manage Subscription",
    "settings.pro.unlock.subtitle": "Unlock the full Bead Studio & photo editor",
    "settings.pro.unlock.title": "Unlock MiLens Pro",
    "settings.section.about": "About",
    "settings.section.appearance": "Appearance",
    "settings.section.notifications": "Notifications",
    "settings.section.privacy": "Data & Privacy",
    "settings.section.pro": "MiLens Pro",
    "settings.section.support": "Support",
    "settings.support.help": "Help",
    "tab.create": "Create",
    "tab.home": "Home",
    "tab.pets": "Pets",
    "tab.settings": "Me",
    # 字面量 key（key 即中文原文）
    "+": "+",
    "\u00b7": "\u00b7",
    "\u2212": "\u2212",
    "AI 模型未就绪，暂不能注册视觉特征": "AI models aren't ready yet \u2014 visual features can't be registered right now.",
    "为它创建第一份档案": "Create its first profile",
    "为它建立一份档案，\n照片、纪念日和故事都会留在这里。": "Create a profile for it \u2014\nphotos, anniversaries, and stories all live here.",
    "从相册里选几张和它有关的照片，\n我们会把每一天整理好。": "Pick a few photos of it from your library,\nand we'll organize every day for you.",
    "你的照片只属于你": "Your photos belong to you",
    "保存": "Save",
    "保存后将以 PNG 格式保留透明背景": "Transparent background will be kept as PNG after saving",
    "先到相册导入照片，再回来生成拼豆图纸": "Import photos from your library first, then come back to create a bead pattern",
    "先留下一张照片": "Add a photo first",
    "分享": "Share",
    "原图预览 · 生成后这里会实时显示拼豆效果": "Original preview \u00b7 the bead effect will appear here after generation",
    "只会从咪Lens 的整理记录中移除，不会删除系统相册原图。": "This only removes it from MiLens' organized library \u2014 originals in your system photo library won't be deleted.",
    "可在「设置 → 隐私 → 照片」中开启，也可以稍后在相册页授权": "You can enable it in Settings \u2192 Privacy \u2192 Photos, or authorize later from the album page",
    "启用\"只拼主体\"时会自动抠图；失败后仍会使用原图继续生成": "Enabling \u201cSubject Only\u201d automatically cuts out the subject; if that fails, generation continues with the original photo",
    "和它一起的每一天": "Every day with them",
    "咪Lens": "MiLens",
    "图纸尺寸": "Pattern Size",
    "头像": "Avatar",
    "头像选择功能将在后续版本支持": "Avatar selection will be available in a future update",
    "字号": "Font Size",
    "它叫什么名字？": "What's its name?",
    "宠物卡片": "Pet Card",
    "将清除 MiLens 中的相册记录、宠物档案及沙盒中已保存的照片副本（导入/编辑产物），且无法恢复。系统相册中的原图不会被删除。": "This clears MiLens' album records, pet profiles, and saved photo copies in the sandbox (imported/edited works). This cannot be undone. Originals in your system photo library won't be deleted.",
    "开始使用": "Get Started",
    "开始注册": "Start Registering",
    "当前方案": "Current Plan",
    "我们找到了它": "We found them",
    "我已阅读并同意": "I have read and agree to",
    "打开本地相册记录时出现问题。您可以重试，或重建本地数据：清除 MiLens 记录，同时也会删除沙盒中已保存的照片副本（导入/编辑产物）。系统相册中的原图不受影响。": "There was a problem opening your local library. You can retry, or rebuild local data: this clears MiLens' records and also deletes photo copies saved in the sandbox (imported/edited works). Originals in your system photo library are unaffected.",
    "扫描系统相册，自动发现你的宠物照片": "Scan your photo library to automatically find pet photos",
    "把它的一生，留在这里": "Keep their whole life here",
    "拼豆图纸": "Bead Pattern",
    "推荐：拼豆插画": "Recommended: Bead Illustration",
    "未标注日期": "No date",
    "本地数据无法加载": "Can't Load Local Data",
    "材料清单": "Materials List",
    "档案加载失败": "Failed to Load Profile",
    "档案已创建": "Profile Created",
    "概括度": "Abstraction",
    "没有符合条件的照片": "No matching photos",
    "添加伙伴": "Add a Companion",
    "添加宠物档案和照片后，这里会自动生成成长时间线": "After you add a pet profile and photos, a growth timeline will appear here automatically",
    "添加照片": "Add Photos",
    "特别的缘分": "What a special connection",
    "知道啦": "Got it",
    "视觉特征": "Visual Features",
    "还没有伙伴档案": "No companion profiles yet",
    "还没有成长记录": "No growth records yet",
    "还没有照片": "No photos yet",
    "这位可爱的宝贝和本APP的开发者同一天出生。\n感谢屏幕前的你，呵护着如此珍贵的生命。": "This lovely one shares a birthday with the app's developer.\nThank you for caring for such a precious life.",
    "选择效果": "Choose Effect",
    "选择照片生成拼豆图纸": "Pick a photo to create a bead pattern",
    "重要事件": "Important Events",
    "随时可以在相册页扫描并导入更多照片": "You can scan and import more photos from the album page anytime",
    "隐私优先": "Privacy First",
    "颜色数量": "Color Count",
    "颜色过渡": "Color Transition",
    "\U0001F382": "\U0001F382",
    "💡 放大图纸即可显示编号": "\U0001F4A1 Zoom into the pattern to reveal numbers",
}

# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #

def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for path in XCSTRINGS:
        if not path.exists():
            sys.exit(f"错误：找不到 {path}")
        sheet_from_xcstrings(wb, path)

    add_overview(wb)
    add_metadata(wb)
    add_subscriptions(wb)
    add_privacy(wb)
    add_screenshots(wb)
    add_aso(wb)
    add_glossary(wb)

    wb.save(OUT)
    print(f"已生成全球本地化资产工作簿 -> {OUT}")
    print(f"  Localizable/InfoPlist 由 localization.py 工具链格式生成；"
          f"en 初译 {len(EN_TRANSLATIONS)} 条（state=needs_review）")
    print("  翻译完成后：python tools/localization.py import build/... --lang en")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
